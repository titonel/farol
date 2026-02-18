from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Q
from functools import wraps
import csv
import io
import re
from datetime import date
from decimal import Decimal, InvalidOperation

from .forms import (
    LoginForm, TrocaSenhaForm, UsuarioForm, EmpresaForm, MedicoForm,
    CirurgiaForm, CirurgiaUploadForm, ExameForm, ServicoMedicoForm,
    ProducaoUploadForm,
)
from .models import Usuario, Empresa, Medico, Cirurgia, Exame, ServicoMedico, ProducaoMensal


# DECORATOR PARA TIER 5
def tier5_required(view_func):
    """Decorator que restringe acesso apenas para usuários Tier 5."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if request.user.primeiro_acesso:
            return redirect('trocar_senha')
        if not request.user.is_admin():
            messages.error(request, 'Acesso negado. Esta área é restrita a administradores (Tier 5).')
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


def login_view(request):
    """View de login."""
    if request.user.is_authenticated:
        if request.user.primeiro_acesso:
            return redirect('trocar_senha')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            
            if user is not None:
                login(request, user)
                if user.primeiro_acesso:
                    messages.warning(request, 'Você precisa trocar sua senha antes de continuar.')
                    return redirect('trocar_senha')
                messages.success(request, f'Bem-vindo(a), {user.nome_completo}!')
                return redirect('dashboard')
            else:
                messages.error(request, 'Usuário ou senha inválidos.')
    else:
        form = LoginForm()
    
    return render(request, 'core/login.html', {'form': form})


@login_required
def logout_view(request):
    """View de logout."""
    logout(request)
    messages.info(request, 'Você saiu do sistema.')
    return redirect('login')


@login_required
def trocar_senha_view(request):
    """View para troca de senha no primeiro acesso."""
    if request.method == 'POST':
        form = TrocaSenhaForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            user.primeiro_acesso = False
            user.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Senha alterada com sucesso!')
            return redirect('dashboard')
    else:
        form = TrocaSenhaForm(request.user)
    
    return render(request, 'core/trocar_senha.html', {'form': form})


@login_required
def dashboard_view(request):
    """Landing page após login."""
    if request.user.primeiro_acesso:
        return redirect('trocar_senha')
    
    context = {
        'total_usuarios': Usuario.objects.count(),
        'total_empresas': Empresa.objects.filter(ativa=True).count(),
        'total_medicos': Medico.objects.filter(ativo=True).count(),
        'total_cirurgias': Cirurgia.objects.filter(ativa=True).count(),
        'total_exames': Exame.objects.filter(ativo=True).count(),
        'total_servicos': ServicoMedico.objects.filter(ativo=True).count(),
    }
    return render(request, 'core/dashboard.html', context)


# CADASTROS

@login_required
def cadastro_menu_view(request):
    """Menu de cadastros."""
    if request.user.primeiro_acesso:
        return redirect('trocar_senha')
    return render(request, 'core/cadastro_menu.html')


# USUARIOS

@login_required
def usuario_lista_view(request):
    """Lista todos os usuários."""
    if not request.user.pode_cadastrar_usuarios():
        messages.error(request, 'Você não tem permissão para acessar esta página.')
        return redirect('dashboard')
    
    usuarios = Usuario.objects.all().order_by('-data_cadastro')
    return render(request, 'core/usuario_lista.html', {'usuarios': usuarios})


@login_required
def usuario_criar_view(request):
    """Cria novo usuário."""
    if not request.user.pode_cadastrar_usuarios():
        messages.error(request, 'Você não tem permissão para cadastrar usuários.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = UsuarioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuário cadastrado com sucesso!')
            return redirect('usuario_lista')
    else:
        form = UsuarioForm()
    
    return render(request, 'core/usuario_form.html', {'form': form, 'acao': 'Cadastrar'})


# EMPRESAS

@login_required
def empresa_lista_view(request):
    """Lista todas as empresas."""
    empresas = Empresa.objects.all().order_by('razao_social')
    return render(request, 'core/empresa_lista.html', {'empresas': empresas})


@login_required
def empresa_criar_view(request):
    """Cria nova empresa."""
    if request.method == 'POST':
        form = EmpresaForm(request.POST)
        if form.is_valid():
            empresa = form.save(commit=False)
            empresa.cadastrado_por = request.user
            empresa.save()
            messages.success(request, 'Empresa cadastrada com sucesso!')
            return redirect('empresa_lista')
    else:
        form = EmpresaForm()
    
    return render(request, 'core/empresa_form.html', {'form': form, 'acao': 'Cadastrar'})


@login_required
def empresa_editar_view(request, pk):
    """Edita uma empresa existente."""
    empresa = get_object_or_404(Empresa, pk=pk)
    
    if request.method == 'POST':
        form = EmpresaForm(request.POST, instance=empresa)
        if form.is_valid():
            form.save()
            messages.success(request, 'Empresa atualizada com sucesso!')
            return redirect('empresa_lista')
    else:
        form = EmpresaForm(instance=empresa)
    
    return render(request, 'core/empresa_form.html', {'form': form, 'acao': 'Editar'})


# MEDICOS

@login_required
def medico_lista_view(request):
    """Lista todos os médicos."""
    medicos = Medico.objects.all().order_by('nome_completo')
    return render(request, 'core/medico_lista.html', {'medicos': medicos})


@login_required
def medico_criar_view(request):
    """Cria novo médico."""
    if request.method == 'POST':
        form = MedicoForm(request.POST)
        if form.is_valid():
            medico = form.save(commit=False)
            medico.cadastrado_por = request.user
            medico.save()
            messages.success(request, 'Médico cadastrado com sucesso!')
            return redirect('medico_lista')
    else:
        form = MedicoForm()
    
    return render(request, 'core/medico_form.html', {'form': form, 'acao': 'Cadastrar'})


@login_required
def medico_editar_view(request, pk):
    """Edita um médico existente."""
    medico = get_object_or_404(Medico, pk=pk)
    
    if request.method == 'POST':
        form = MedicoForm(request.POST, instance=medico)
        if form.is_valid():
            form.save()
            messages.success(request, 'Médico atualizado com sucesso!')
            return redirect('medico_lista')
    else:
        form = MedicoForm(instance=medico)
    
    return render(request, 'core/medico_form.html', {'form': form, 'acao': 'Editar'})


# ===== ÁREA ADMINISTRATIVA (TIER 5) =====

@tier5_required
def admin_menu_view(request):
    """Menu da área administrativa."""
    context = {
        'total_cirurgias': Cirurgia.objects.count(),
        'total_exames': Exame.objects.count(),
        'total_servicos': ServicoMedico.objects.count(),
    }
    return render(request, 'core/admin/menu.html', context)


# CIRURGIAS

@tier5_required
def cirurgia_lista_view(request):
    """Lista todas as cirurgias."""
    cirurgias = Cirurgia.objects.all().order_by('especialidade', 'descricao')
    return render(request, 'core/admin/cirurgia_lista.html', {'cirurgias': cirurgias})


@tier5_required
def cirurgia_criar_view(request):
    """Cria nova cirurgia."""
    if request.method == 'POST':
        form = CirurgiaForm(request.POST)
        if form.is_valid():
            cirurgia = form.save(commit=False)
            cirurgia.cadastrado_por = request.user
            cirurgia.save()
            messages.success(request, 'Cirurgia cadastrada com sucesso!')
            return redirect('cirurgia_lista')
    else:
        form = CirurgiaForm()
    
    return render(request, 'core/admin/cirurgia_form.html', {'form': form, 'acao': 'Cadastrar'})


@tier5_required
def cirurgia_editar_view(request, pk):
    """Edita uma cirurgia existente."""
    cirurgia = get_object_or_404(Cirurgia, pk=pk)
    
    if request.method == 'POST':
        form = CirurgiaForm(request.POST, instance=cirurgia)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cirurgia atualizada com sucesso!')
            return redirect('cirurgia_lista')
    else:
        form = CirurgiaForm(instance=cirurgia)
    
    return render(request, 'core/admin/cirurgia_form.html', {'form': form, 'acao': 'Editar'})


@tier5_required
def cirurgia_upload_view(request):
    """Upload de CSV de cirurgias."""
    if request.method == 'POST':
        form = CirurgiaUploadForm(request.POST, request.FILES)
        if form.is_valid():
            arquivo = request.FILES['arquivo_csv']
            
            # Lê o arquivo CSV
            arquivo.seek(0)
            decoded_file = arquivo.read().decode('utf-8-sig')  # utf-8-sig remove BOM automaticamente
            csv_file = io.StringIO(decoded_file)
            
            # Detecta automaticamente o delimitador (vírgula ou ponto e vírgula)
            sample = csv_file.read(1024)
            csv_file.seek(0)
            sniffer = csv.Sniffer()
            try:
                delimiter = sniffer.sniff(sample).delimiter
            except:
                delimiter = ';'  # Default para ponto e vírgula se não detectar
            
            reader = csv.DictReader(csv_file, delimiter=delimiter)
            
            # Normaliza os nomes das colunas (remove espaços extras)
            if reader.fieldnames:
                reader.fieldnames = [field.strip() for field in reader.fieldnames]
            
            # Mapeia variações de nomes de colunas
            def get_column(row, variations):
                for var in variations:
                    # Procura exatamente
                    if var in row:
                        value = row[var]
                        # Remove espaços e ponto e vírgula final
                        if value:
                            return value.strip().rstrip(';')
                    # Tenta com espaços substituídos por underscore
                    var_underscore = var.replace(' ', '_')
                    if var_underscore in row:
                        value = row[var_underscore]
                        if value:
                            return value.strip().rstrip(';')
                return None
            
            sucesso = 0
            erro = 0
            erros_detalhados = []
            linhas_processadas = 0
            
            for i, row in enumerate(reader, start=2):  # Começa do 2 (header é linha 1)
                try:
                    # Pula linhas vazias
                    if not any(row.values()):
                        continue
                    
                    linhas_processadas += 1
                    
                    codigo = get_column(row, ['Codigo SIGTAP', 'codigo_sigtap', 'codigo'])
                    descricao = get_column(row, ['Descricao', 'descricao'])
                    valor_str = get_column(row, ['Valor', 'valor'])
                    tipo = get_column(row, ['Tipo Cirurgia', 'tipo_cirurgia', 'tipo'])
                    especialidade = get_column(row, ['Especialidade', 'especialidade'])
                    
                    # Valida dados obrigatórios
                    if not codigo:
                        erros_detalhados.append(f"Linha {i}: Código SIGTAP ausente")
                        erro += 1
                        continue
                    
                    if not descricao:
                        erros_detalhados.append(f"Linha {i}: Descrição ausente")
                        erro += 1
                        continue
                    
                    if not tipo:
                        erros_detalhados.append(f"Linha {i}: Tipo cirurgia ausente")
                        erro += 1
                        continue
                    
                    # Valor padrão 0 se não informado
                    if not valor_str:
                        valor = Decimal('0.00')
                    else:
                        try:
                            valor = Decimal(valor_str.replace(',', '.'))
                        except:
                            erros_detalhados.append(f"Linha {i}: Valor inválido '{valor_str}'")
                            erro += 1
                            continue
                    
                    # Especialidade padrão se não informada
                    if not especialidade:
                        especialidade = 'Não especificada'
                    
                    # Mapeia tipo de cirurgia
                    tipo_stripped = tipo.strip()
                    if tipo_stripped.upper() == 'CMA':
                        tipo_cirurgia = 'CMA'
                    elif tipo_stripped.lower() == 'cma':
                        tipo_cirurgia = 'cma'
                    else:
                        tipo_upper = tipo_stripped.upper()
                        if 'MAIOR' in tipo_upper or tipo_upper == 'CMA':
                            tipo_cirurgia = 'CMA'
                        elif 'MENOR' in tipo_upper:
                            tipo_cirurgia = 'cma'
                        else:
                            erros_detalhados.append(f"Linha {i}: Tipo inválido '{tipo}'. Use 'CMA' ou 'cma'")
                            erro += 1
                            continue
                    
                    # Cria ou atualiza cirurgia
                    cirurgia, created = Cirurgia.objects.update_or_create(
                        codigo_sigtap=codigo.strip(),
                        defaults={
                            'descricao': descricao.strip(),
                            'valor': valor,
                            'tipo_cirurgia': tipo_cirurgia,
                            'especialidade': especialidade.strip(),
                            'cadastrado_por': request.user,
                        }
                    )
                    sucesso += 1
                    
                except Exception as e:
                    erros_detalhados.append(f"Linha {i}: {str(e)}")
                    erro += 1
            
            # Mensagens de resultado
            if linhas_processadas == 0:
                messages.error(request, 'Arquivo CSV vazio ou sem dados válidos.')
            elif sucesso > 0:
                messages.success(request, f'{sucesso} cirurgia(s) importada(s) com sucesso!')
            
            if erro > 0:
                mensagem_erros = f'{erro} linha(s) com erro.'
                if erros_detalhados:
                    mensagem_erros += f' Primeiros erros: {"; ".join(erros_detalhados[:5])}'
                messages.warning(request, mensagem_erros)
            
            return redirect('cirurgia_lista')
    else:
        form = CirurgiaUploadForm()
    
    return render(request, 'core/admin/cirurgia_upload.html', {'form': form})


# EXAMES

@tier5_required
def exame_lista_view(request):
    """Lista todos os exames."""
    exames = Exame.objects.all().order_by('tipo_exame', 'descricao')
    return render(request, 'core/admin/exame_lista.html', {'exames': exames})


@tier5_required
def exame_criar_view(request):
    """Cria novo exame."""
    if request.method == 'POST':
        form = ExameForm(request.POST)
        if form.is_valid():
            exame = form.save(commit=False)
            exame.cadastrado_por = request.user
            exame.save()
            messages.success(request, 'Exame cadastrado com sucesso!')
            return redirect('exame_lista')
    else:
        form = ExameForm()
    
    return render(request, 'core/admin/exame_form.html', {'form': form, 'acao': 'Cadastrar'})


@tier5_required
def exame_editar_view(request, pk):
    """Edita um exame existente."""
    exame = get_object_or_404(Exame, pk=pk)
    
    if request.method == 'POST':
        form = ExameForm(request.POST, instance=exame)
        if form.is_valid():
            form.save()
            messages.success(request, 'Exame atualizado com sucesso!')
            return redirect('exame_lista')
    else:
        form = ExameForm(instance=exame)
    
    return render(request, 'core/admin/exame_form.html', {'form': form, 'acao': 'Editar'})


# SERVIÇOS MÉDICOS

@tier5_required
def servico_lista_view(request):
    """Lista todos os serviços médicos."""
    servicos = ServicoMedico.objects.all().order_by('especialidade', 'descricao')
    return render(request, 'core/admin/servico_lista.html', {'servicos': servicos})


@tier5_required
def servico_criar_view(request):
    """Cria novo serviço médico."""
    if request.method == 'POST':
        form = ServicoMedicoForm(request.POST)
        if form.is_valid():
            servico = form.save(commit=False)
            servico.cadastrado_por = request.user
            servico.save()
            messages.success(request, 'Serviço médico cadastrado com sucesso!')
            return redirect('servico_lista')
    else:
        form = ServicoMedicoForm()
    
    return render(request, 'core/admin/servico_form.html', {'form': form, 'acao': 'Cadastrar'})


@tier5_required
def servico_editar_view(request, pk):
    """Edita um serviço médico existente."""
    servico = get_object_or_404(ServicoMedico, pk=pk)
    
    if request.method == 'POST':
        form = ServicoMedicoForm(request.POST, instance=servico)
        if form.is_valid():
            form.save()
            messages.success(request, 'Serviço médico atualizado com sucesso!')
            return redirect('servico_lista')
    else:
        form = ServicoMedicoForm(instance=servico)
    
    return render(request, 'core/admin/servico_form.html', {'form': form, 'acao': 'Editar'})


# ===== MÓDULO DE PRODUÇÃO =====

_MESES_PT = {
    'janeiro': 1, 'fevereiro': 2, 'marco': 3, 'março': 3,
    'abril': 4, 'maio': 5, 'junho': 6,
    'julho': 7, 'agosto': 8, 'setembro': 9,
    'outubro': 10, 'novembro': 11, 'dezembro': 12,
    'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
    'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12,
}

_NOMES_MESES = {
    1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
    5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
    9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro',
}


def _parse_mes_ano(cell_value):
    """Interpreta o conteúdo da célula F3 e retorna o primeiro dia do mês."""
    if cell_value is None:
        raise ValueError("Célula F3 está vazia.")

    if hasattr(cell_value, 'year') and hasattr(cell_value, 'month'):
        return date(cell_value.year, cell_value.month, 1)

    value = str(cell_value).strip()

    m = re.match(r'^(\d{1,2})[/\-](\d{4})$', value)
    if m:
        return date(int(m.group(2)), int(m.group(1)), 1)

    m = re.match(r'^(\d{4})[/\-](\d{1,2})$', value)
    if m:
        return date(int(m.group(1)), int(m.group(2)), 1)

    m = re.match(r'^([A-Za-zÀ-ÿ]+)[/ ](\d{4})$', value)
    if m:
        mes_str = m.group(1).lower()
        mes_str = mes_str.translate(str.maketrans('áéíóúâêîôûãõç', 'aeiouaeiouaoc'))
        mes = _MESES_PT.get(mes_str)
        if mes:
            return date(int(m.group(2)), mes, 1)

    raise ValueError(
        f"Não foi possível interpretar a data '{cell_value}' da célula F3. "
        "Use formatos como: MM/AAAA, Janeiro/AAAA ou Janeiro AAAA."
    )


def _to_int(v):
    if v is None or (isinstance(v, str) and v.strip() == ''):
        return None
    try:
        return int(float(str(v).replace(',', '.')))
    except (ValueError, TypeError):
        return None


def _to_decimal_str(v):
    if v is None or (isinstance(v, str) and v.strip() == ''):
        return None
    try:
        s = str(v).replace(',', '.').replace('%', '').strip()
        return str(Decimal(s))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _parse_xlsx(arquivo):
    """Lê arquivo .xlsx e retorna (mes_ano, lista_de_registros)."""
    import openpyxl
    wb = openpyxl.load_workbook(arquivo, data_only=True)
    ws = wb.active

    cell_f3 = ws.cell(row=3, column=6).value
    mes_ano = _parse_mes_ano(cell_f3)

    registros = []
    for row_idx in range(8, ws.max_row + 1):
        vals = [ws.cell(row=row_idx, column=c).value for c in range(1, 14)]
        if all(v is None or str(v).strip() == '' for v in vals):
            continue
        especialidade = vals[0]
        if especialidade is None or str(especialidade).strip() == '':
            continue
        registros.append({
            'especialidade': str(especialidade).strip(),
            'vagas_ofertadas': _to_int(vals[1]),
            'total_agendamentos': _to_int(vals[2]),
            'perc_agendamentos': _to_decimal_str(vals[3]),
            'agendamentos_cota': _to_int(vals[4]),
            'perc_cota': _to_decimal_str(vals[5]),
            'vagas_bolsao': _to_int(vals[6]),
            'perc_bolsao': _to_decimal_str(vals[7]),
            'vagas_nao_distribuidas': _to_int(vals[8]),
            'perc_nao_distribuidas': _to_decimal_str(vals[9]),
            'vagas_extras': _to_int(vals[10]),
            'perc_extras': _to_decimal_str(vals[11]),
            'perc_desperdicadas': _to_decimal_str(vals[12]),
        })
    return mes_ano, registros


def _parse_xls(arquivo):
    """Lê arquivo .xls e retorna (mes_ano, lista_de_registros)."""
    import xlrd
    content = arquivo.read()
    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)

    cell_f3_raw = ws.cell(2, 5)
    if cell_f3_raw.ctype == xlrd.XL_CELL_DATE:
        dt = xlrd.xldate_as_tuple(cell_f3_raw.value, wb.datemode)
        mes_ano = date(dt[0], dt[1], 1)
    else:
        mes_ano = _parse_mes_ano(cell_f3_raw.value)

    registros = []
    for row_idx in range(7, ws.nrows):
        vals = [ws.cell(row_idx, c).value for c in range(13)]
        if all(v == '' or v is None for v in vals):
            continue
        especialidade = vals[0]
        if not especialidade or str(especialidade).strip() == '':
            continue
        registros.append({
            'especialidade': str(especialidade).strip(),
            'vagas_ofertadas': _to_int(vals[1]),
            'total_agendamentos': _to_int(vals[2]),
            'perc_agendamentos': _to_decimal_str(vals[3]),
            'agendamentos_cota': _to_int(vals[4]),
            'perc_cota': _to_decimal_str(vals[5]),
            'vagas_bolsao': _to_int(vals[6]),
            'perc_bolsao': _to_decimal_str(vals[7]),
            'vagas_nao_distribuidas': _to_int(vals[8]),
            'perc_nao_distribuidas': _to_decimal_str(vals[9]),
            'vagas_extras': _to_int(vals[10]),
            'perc_extras': _to_decimal_str(vals[11]),
            'perc_desperdicadas': _to_decimal_str(vals[12]),
        })
    return mes_ano, registros


@login_required
def producao_menu_view(request):
    """Landing page do módulo de produção."""
    if request.user.primeiro_acesso:
        return redirect('trocar_senha')
    return render(request, 'core/producao_menu.html')


@login_required
def producao_upload_view(request):
    """Upload da planilha de produção mensal."""
    if request.user.primeiro_acesso:
        return redirect('trocar_senha')

    if request.method == 'POST':
        form = ProducaoUploadForm(request.POST, request.FILES)
        if form.is_valid():
            arquivo = request.FILES['arquivo']
            nome = arquivo.name.lower()
            try:
                if nome.endswith('.xlsx'):
                    mes_ano, registros = _parse_xlsx(arquivo)
                else:
                    mes_ano, registros = _parse_xls(arquivo)

                if not registros:
                    messages.error(request, 'Nenhum dado encontrado no arquivo. Verifique a estrutura da planilha.')
                    return render(request, 'core/producao_upload.html', {'form': form})

                request.session['producao_upload'] = {
                    'mes_ano': mes_ano.isoformat(),
                    'mes_ano_display': f"{_NOMES_MESES[mes_ano.month]}/{mes_ano.year}",
                    'registros': registros,
                }
                return redirect('producao_confirmar')

            except ValueError as e:
                messages.error(request, f'Erro ao ler o arquivo: {e}')
            except Exception as e:
                messages.error(request, f'Erro inesperado ao processar o arquivo: {e}')
    else:
        form = ProducaoUploadForm()

    return render(request, 'core/producao_upload.html', {'form': form})


@login_required
def producao_confirmar_view(request):
    """Exibe os dados carregados para confirmação antes de gravar."""
    if request.user.primeiro_acesso:
        return redirect('trocar_senha')

    dados = request.session.get('producao_upload')
    if not dados:
        messages.warning(request, 'Sessão expirada. Faça o upload novamente.')
        return redirect('producao_upload')

    mes_ano = date.fromisoformat(dados['mes_ano'])
    mes_ano_display = dados.get('mes_ano_display', mes_ano.strftime('%m/%Y'))
    registros = dados['registros']
    existe = ProducaoMensal.objects.filter(mes_ano=mes_ano).exists()

    if request.method == 'POST':
        try:
            with transaction.atomic():
                ProducaoMensal.objects.filter(mes_ano=mes_ano).delete()
                for reg in registros:
                    def _d(v):
                        return Decimal(v) if v is not None else None
                    ProducaoMensal.objects.create(
                        mes_ano=mes_ano,
                        especialidade=reg['especialidade'],
                        vagas_ofertadas=reg['vagas_ofertadas'],
                        total_agendamentos=reg['total_agendamentos'],
                        perc_agendamentos=_d(reg['perc_agendamentos']),
                        agendamentos_cota=reg['agendamentos_cota'],
                        perc_cota=_d(reg['perc_cota']),
                        vagas_bolsao=reg['vagas_bolsao'],
                        perc_bolsao=_d(reg['perc_bolsao']),
                        vagas_nao_distribuidas=reg['vagas_nao_distribuidas'],
                        perc_nao_distribuidas=_d(reg['perc_nao_distribuidas']),
                        vagas_extras=reg['vagas_extras'],
                        perc_extras=_d(reg['perc_extras']),
                        perc_desperdicadas=_d(reg['perc_desperdicadas']),
                        importado_por=request.user,
                    )
            del request.session['producao_upload']
            messages.success(
                request,
                f'{len(registros)} registro(s) gravado(s) com sucesso para {mes_ano_display}!'
            )
            return redirect('producao_dashboard')
        except Exception as e:
            messages.error(request, f'Erro ao gravar os dados: {e}')

    context = {
        'mes_ano': mes_ano,
        'mes_ano_display': mes_ano_display,
        'registros': registros,
        'existe': existe,
        'total': len(registros),
    }
    return render(request, 'core/producao_confirmar.html', context)


@login_required
def producao_dashboard_view(request):
    """Dashboard de acompanhamento da produção mensal."""
    if request.user.primeiro_acesso:
        return redirect('trocar_senha')

    meses_disponiveis = (
        ProducaoMensal.objects
        .values_list('mes_ano', flat=True)
        .distinct()
        .order_by('-mes_ano')
    )

    mes_selecionado = None
    mes_selecionado_str = request.GET.get('mes')
    if mes_selecionado_str:
        try:
            mes_selecionado = date.fromisoformat(mes_selecionado_str)
        except ValueError:
            pass

    if mes_selecionado is None and meses_disponiveis:
        mes_selecionado = meses_disponiveis[0]

    producoes = None
    mes_selecionado_display = None
    if mes_selecionado:
        producoes = ProducaoMensal.objects.filter(mes_ano=mes_selecionado).order_by('especialidade')
        mes_selecionado_display = f"{_NOMES_MESES[mes_selecionado.month]}/{mes_selecionado.year}"

    meses_com_display = [
        {'valor': m.isoformat(), 'display': f"{_NOMES_MESES[m.month]}/{m.year}"}
        for m in meses_disponiveis
    ]

    context = {
        'meses_com_display': meses_com_display,
        'mes_selecionado': mes_selecionado,
        'mes_selecionado_display': mes_selecionado_display,
        'producoes': producoes,
    }
    return render(request, 'core/producao_dashboard.html', context)
