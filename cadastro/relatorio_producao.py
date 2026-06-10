"""Gerador do Relatório de Escala de Produção Médica.

Uso:
    from cadastro.relatorio_producao import criar_relatorio
    wb = criar_relatorio(mes_ini=1, ano_ini=2026, nome_empresa='...', ...)
    wb.save('relatorio.xlsx')
"""
from datetime import date, timedelta
import calendar

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paleta extraída do modelo original (DR-TATIANA-ROZOV.XLS) ────────────────
TEAL       = "33CCCC"   # cabeçalhos principais
BLUE_LIGHT = "99CCFF"   # dias do calendário
GRAY_LIGHT = "C0C0C0"   # linhas de dados
DARK_GRAY  = "333333"   # texto
WHITE      = "FFFFFF"

DIAS_PT = {0: "S", 1: "T", 2: "Q", 3: "Q", 4: "S", 5: "S", 6: "D"}

MESES_PT = [
    "janeiro", "fevereiro", "março", "abril", "maio", "junho",
    "julho", "agosto", "setembro", "outubro", "novembro", "dezembro",
]

# ── Helpers de estilo ────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=11, color=DARK_GRAY):
    return Font(name="Arial", bold=bold, size=size, color=color)

def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def _apply(cell, fnt=None, fll=None, aln=None):
    brd = _border()
    if fnt:
        cell.font = fnt
    if fll:
        cell.fill = fll
    if aln:
        cell.alignment = aln
    cell.border = brd


def _gerar_periodo(mes_ini, ano_ini):
    """Lista de (dia, mes, ano, abrev_dia_semana) do dia 21 ao dia 20 do mês seguinte."""
    dias = []
    data = date(ano_ini, mes_ini, 21)
    while True:
        dias.append((data.day, data.month, data.year, DIAS_PT[data.weekday()]))
        if data.day == 20 and data.month != mes_ini:
            break
        data += timedelta(days=1)
    return dias


def _nome_mes(mes, ano):
    return f"{MESES_PT[mes - 1].upper()}/{ano}"


def _data_pt(d):
    return (
        f"AME Caraguatatuba, {d.day} de {MESES_PT[d.month - 1]} de {d.year}"
    )


# ── Função principal ─────────────────────────────────────────────────────────
def criar_relatorio(mes_ini, ano_ini, nome_empresa, especialidade,
                    servicos, prestador_nome, crm, observacoes=""):
    """
    Parâmetros
    ----------
    mes_ini      : int  – mês de início do período (1–12)
    ano_ini      : int  – ano de início
    nome_empresa : str  – razão social do prestador
    especialidade: str  – especialidade médica
    servicos     : list[dict] com chaves:
                     descricao   : str
                     cod         : int
                     agenda      : str
                     estimativa  : int   (qtd estimada em contrato)
                     valor_unit  : float
                     producao    : dict  { (dia, mes): qtd }  — dias com produção registrada
    prestador_nome: str
    crm          : str
    observacoes  : str  (opcional)

    Retorna
    -------
    openpyxl.Workbook
    """
    dias = _gerar_periodo(mes_ini, ano_ini)
    mes_fim = 1 if mes_ini == 12 else mes_ini + 1
    ano_fim = ano_ini + 1 if mes_ini == 12 else ano_ini
    n_dias = len(dias)
    n_srv = len(servicos)

    idx_mes2 = next(i for i, d in enumerate(dias) if d[0] == 1)

    COL_DIA1 = 7
    COL_SUMBASE = COL_DIA1 + n_dias

    wb = Workbook()
    ws = wb.active
    ws.title = f"{_nome_mes(mes_ini, ano_ini)[:3]}-{_nome_mes(mes_fim, ano_fim)[:3]}"

    # ── Larguras ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 3
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 5
    for i in range(n_dias):
        ws.column_dimensions[get_column_letter(COL_DIA1 + i)].width = 4
    for i in range(n_srv * 2 + 18):
        ws.column_dimensions[get_column_letter(COL_SUMBASE + i)].width = 13

    # ── Alturas ───────────────────────────────────────────────────────────
    for r, h in [(4, 22), (5, 18), (6, 18), (8, 18), (9, 18), (10, 14)]:
        ws.row_dimensions[r].height = h

    last_info_col = get_column_letter(COL_SUMBASE - 1)

    # ── Título (row 4) ────────────────────────────────────────────────────
    ws.merge_cells("C4:F4")
    c = ws["C4"]
    c.value = "ESCALA DE PRODUÇÃO MÉDICA"
    c.font = _font(bold=True, size=14)
    c.alignment = _align("left")

    ws.merge_cells(f"G4:{last_info_col}4")
    c2 = ws["G4"]
    c2.value = especialidade
    c2.font = _font(size=12)
    c2.alignment = _align("left")

    # ── Empresa (row 5) ───────────────────────────────────────────────────
    ws.merge_cells(f"C5:{last_info_col}5")
    c = ws["C5"]
    c.value = f"Empresa: {nome_empresa}"
    c.font = _font(size=12)
    c.alignment = _align("left")

    # ── Período (row 6) ───────────────────────────────────────────────────
    ws.merge_cells(f"C6:{last_info_col}6")
    c = ws["C6"]
    c.value = (f"Período: 21/{mes_ini:02d}/{ano_ini} a "
               f"20/{mes_fim:02d}/{ano_fim}")
    c.font = _font(size=12)
    c.alignment = _align("left")

    # ── Cabeçalho superior (row 8) ────────────────────────────────────────
    ws.merge_cells("A8:F8")
    c = ws["A8"]
    c.value = "PERÍODO"
    _apply(c, _font(bold=True, size=12, color=WHITE), _fill(TEAL), _align())

    col_m1_end = get_column_letter(COL_DIA1 + idx_mes2 - 1)
    ws.merge_cells(f"{get_column_letter(COL_DIA1)}8:{col_m1_end}8")
    c = ws.cell(8, COL_DIA1)
    c.value = _nome_mes(mes_ini, ano_ini)
    _apply(c, _font(bold=True, size=12, color=WHITE), _fill(TEAL), _align())

    col_m2_start = get_column_letter(COL_DIA1 + idx_mes2)
    col_m2_end = get_column_letter(COL_DIA1 + n_dias - 1)
    ws.merge_cells(f"{col_m2_start}8:{col_m2_end}8")
    c = ws.cell(8, COL_DIA1 + idx_mes2)
    c.value = _nome_mes(mes_fim, ano_fim)
    _apply(c, _font(bold=True, size=12, color=WHITE), _fill(TEAL), _align())

    for i, sv in enumerate(servicos):
        col_q = COL_SUMBASE + i * 2
        ws.merge_cells(f"{get_column_letter(col_q)}8:{get_column_letter(col_q+1)}8")
        c = ws.cell(8, col_q)
        c.value = sv["descricao"]
        _apply(c, _font(bold=True, size=11), _fill(TEAL), _align(wrap=True))

    # ── Sub-cabeçalhos (row 9) ────────────────────────────────────────────
    ws.merge_cells("A9:B9")
    for col, txt in [(1, "Nome"), (3, "CRM"), (4, "Serviço"),
                     (5, "Agenda Prevista"), (6, "Cod.")]:
        c = ws.cell(9, col)
        c.value = txt
        _apply(c, _font(bold=True, size=11), _fill(TEAL), _align())

    for i, (dia, mes, ano, dow) in enumerate(dias):
        c = ws.cell(9, COL_DIA1 + i)
        c.value = dia
        _apply(c, _font(bold=True, size=11), _fill(BLUE_LIGHT), _align())

    for i, sv in enumerate(servicos):
        col_q = COL_SUMBASE + i * 2
        for col, txt in [(col_q, "Qtd."), (col_q + 1, "Valor (R$)")]:
            c = ws.cell(9, col)
            c.value = txt
            _apply(c, _font(bold=True, size=11), _fill(TEAL), _align())

    # ── Dia da semana (row 10) ────────────────────────────────────────────
    for i, (dia, mes, ano, dow) in enumerate(dias):
        c = ws.cell(10, COL_DIA1 + i)
        c.value = dow
        _apply(c, _font(bold=True, size=11), _fill(BLUE_LIGHT), _align())

    # ── Linhas de dados (rows 11+) ────────────────────────────────────────
    for s_idx, sv in enumerate(servicos):
        row = 11 + s_idx
        ws.row_dimensions[row].height = 16

        if s_idx == 0:
            ws.merge_cells(f"A{row}:B{row}")
            c = ws.cell(row, 1)
            c.value = prestador_nome
            _apply(c, _font(bold=True, size=11), _fill(GRAY_LIGHT), _align("left"))
            c = ws.cell(row, 3)
            c.value = crm
            _apply(c, _font(size=11), _fill(GRAY_LIGHT), _align())
        else:
            for col in (1, 2, 3):
                _apply(ws.cell(row, col), fll=_fill(GRAY_LIGHT))

        for col, val, alg in [
            (4, sv["descricao"], "center"),
            (5, sv.get("agenda", ""), "center"),
            (6, sv["cod"], "center"),
        ]:
            c = ws.cell(row, col)
            c.value = val
            _apply(c, _font(size=10), _fill(GRAY_LIGHT), _align(alg))

        prod = sv.get("producao", {})
        for i, (dia, mes, ano, dow) in enumerate(dias):
            col = COL_DIA1 + i
            qtd = prod.get((dia, mes), None)
            c = ws.cell(row, col)
            if qtd is not None:
                c.value = qtd
                _apply(c, _font(bold=True, size=12), _fill(BLUE_LIGHT), _align())
                c.number_format = "#,##0"
            else:
                _apply(c, fll=_fill(GRAY_LIGHT), aln=_align())

        col_q = COL_SUMBASE + s_idx * 2
        col_v = col_q + 1
        rng = f"{get_column_letter(COL_DIA1)}{row}:{get_column_letter(COL_DIA1 + n_dias - 1)}{row}"
        c = ws.cell(row, col_q)
        c.value = f"=SUM({rng})"
        c.number_format = "#,##0"
        _apply(c, _font(bold=True, size=11), _fill(GRAY_LIGHT), _align())

        c = ws.cell(row, col_v)
        c.value = f"={get_column_letter(col_q)}{row}*{sv.get('valor_unit', 0)}"
        c.number_format = "R$ #,##0.00"
        _apply(c, _font(bold=True, size=11), _fill(GRAY_LIGHT), _align("left"))

    # ── Linha de total (row 11+n_srv) ─────────────────────────────────────
    row_total = 11 + n_srv
    ws.row_dimensions[row_total].height = 16
    ws.merge_cells(
        f"A{row_total}:{get_column_letter(COL_DIA1 + n_dias - 1)}{row_total}"
    )
    c = ws.cell(row_total, 1)
    c.value = "TOTAL"
    _apply(c, _font(bold=True, size=12, color=WHITE), _fill(TEAL), _align("right"))

    for i in range(n_srv):
        col_q = COL_SUMBASE + i * 2
        col_v = col_q + 1
        c = ws.cell(row_total, col_q)
        c.value = f"=SUM({get_column_letter(col_q)}11:{get_column_letter(col_q)}{row_total - 1})"
        c.number_format = "#,##0"
        _apply(c, _font(bold=True, size=12, color=WHITE), _fill(TEAL), _align())
        c = ws.cell(row_total, col_v)
        c.value = f"=SUM({get_column_letter(col_v)}11:{get_column_letter(col_v)}{row_total - 1})"
        c.number_format = "R$ #,##0.00"
        _apply(c, _font(bold=True, size=12, color=WHITE), _fill(TEAL), _align())

    # ── Legenda ───────────────────────────────────────────────────────────
    row_leg = row_total + 2
    ws.cell(row_leg, 1).value = "OBSERVAÇÕES DO PERÍODO"
    ws.cell(row_leg, 1).font = _font(bold=True, size=11)
    ws.cell(row_leg, 7).value = "LEGENDA:"
    ws.cell(row_leg, 7).font = _font(bold=True, size=11)

    for r in range(row_leg + 1, row_leg + 8):
        for col in range(2, 6):
            ws.cell(r, col).border = _border()
    if observacoes:
        ws.cell(row_leg + 1, 2).value = observacoes
        ws.cell(row_leg + 1, 2).font = _font(size=10)
        ws.cell(row_leg + 1, 2).alignment = _align("left", wrap=True)

    legendas = [
        ("", "Atendimento previsto"),
        ("", "Atendimento previsto e realizado"),
        ("", "Atendimento não previsto mas realizado"),
        ("", "Final de Semana / Feriado"),
        ("F", "Falta"),
        ("AB", "Agenda Bloqueada"),
        ("TP", "Troca de Profissional"),
    ]
    for idx, (cod, desc) in enumerate(legendas):
        r = row_leg + 1 + idx
        ws.cell(r, 6).value = cod
        ws.cell(r, 7).value = desc
        ws.cell(r, 7).font = _font(size=10)
        ws.cell(r, 7).alignment = _align("left")
        if cod == "":
            ws.cell(r, 6).fill = _fill(BLUE_LIGHT if idx < 2 else GRAY_LIGHT)

    # ── Resumo do Serviço em Contrato ─────────────────────────────────────
    col_res = COL_SUMBASE
    ws.merge_cells(
        f"{get_column_letter(col_res)}{row_leg}:"
        f"{get_column_letter(col_res + 15)}{row_leg}"
    )
    c = ws.cell(row_leg, col_res)
    c.value = "RESUMO DO SERVIÇO EM CONTRATO"
    c.font = _font(bold=True, size=11)
    c.alignment = _align()

    row_res_hdr = row_leg + 1
    for col, txt in [
        (col_res, "Serviço"),
        (col_res + 6, "Estimativa em Contrato"),
        (col_res + 11, "Total por Serviço"),
        (col_res + 13, "%"),
        (col_res + 14, "Valor Unitário"),
        (col_res + 15, "Valor Total"),
    ]:
        c = ws.cell(row_res_hdr, col)
        c.value = txt
        _apply(c, _font(bold=True, size=10), _fill(TEAL), _align(wrap=True))

    for i, sv in enumerate(servicos):
        r = row_res_hdr + 1 + i
        ws.merge_cells(
            f"{get_column_letter(col_res)}{r}:"
            f"{get_column_letter(col_res + 5)}{r}"
        )
        c = ws.cell(r, col_res)
        c.value = sv["descricao"]
        _apply(c, _font(size=10), _fill(GRAY_LIGHT), _align())

        est_col = col_res + 6
        c = ws.cell(r, est_col)
        c.value = sv.get("estimativa", 0)
        c.number_format = "#,##0"
        _apply(c, _font(size=10), _fill(GRAY_LIGHT), _align())

        tot_col = col_res + 11
        c = ws.cell(r, tot_col)
        c.value = f"={get_column_letter(COL_SUMBASE + i * 2)}{row_total}"
        c.number_format = "#,##0"
        _apply(c, _font(size=10), _fill(GRAY_LIGHT), _align())

        pct_col = col_res + 13
        c = ws.cell(r, pct_col)
        c.value = (
            f"=IF({get_column_letter(est_col)}{r}=0,0,"
            f"{get_column_letter(tot_col)}{r}/{get_column_letter(est_col)}{r})"
        )
        c.number_format = "0.00%"
        _apply(c, _font(size=10), _fill(GRAY_LIGHT), _align())

        vunit_col = col_res + 14
        c = ws.cell(r, vunit_col)
        c.value = sv.get("valor_unit", 0)
        c.number_format = "R$ #,##0.00"
        _apply(c, _font(size=10), _fill(GRAY_LIGHT), _align())

        vtot_col = col_res + 15
        c = ws.cell(r, vtot_col)
        c.value = f"={get_column_letter(COL_SUMBASE + i * 2 + 1)}{row_total}"
        c.number_format = "R$ #,##0.00"
        _apply(c, _font(size=10), _fill(GRAY_LIGHT), _align())

    r_total_res = row_res_hdr + 1 + n_srv
    ws.row_dimensions[r_total_res].height = 16
    ws.merge_cells(
        f"{get_column_letter(col_res)}{r_total_res}:"
        f"{get_column_letter(col_res + 14)}{r_total_res}"
    )
    c = ws.cell(r_total_res, col_res)
    c.value = "TOTAL GERAL"
    _apply(c, _font(bold=True, size=11, color=WHITE), _fill(TEAL), _align())

    vtot_col = col_res + 15
    c = ws.cell(r_total_res, vtot_col)
    c.value = (
        f"=SUM({get_column_letter(vtot_col)}{row_res_hdr + 1}:"
        f"{get_column_letter(vtot_col)}{r_total_res - 1})"
    )
    c.number_format = "R$ #,##0.00"
    _apply(c, _font(bold=True, size=11, color=WHITE), _fill(TEAL), _align())

    # ── Assinatura ────────────────────────────────────────────────────────
    row_sig = r_total_res + 3
    ws.merge_cells(
        f"{get_column_letter(col_res)}{row_sig}:"
        f"{get_column_letter(col_res + 15)}{row_sig}"
    )
    c = ws.cell(row_sig, col_res)
    c.value = _data_pt(date.today())
    c.font = _font(size=11)
    c.alignment = _align()

    # ── Configurações de página ───────────────────────────────────────────
    ws.freeze_panes = "G11"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    return wb
